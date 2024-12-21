import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from collections import defaultdict
import os

# 파일 경로 전역 변수
input_file = ""
output_file = ""

# 파일 선택 함수
def select_input_file():
    global input_file
    input_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    input_file_label.config(text=f": {input_file}")

def select_output_file():
    global output_file
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    output_file_label.config(text=f": {output_file}")

# 처리 함수
def process_files():
    if not input_file or not output_file:
        messagebox.showerror("파일 오류", "입력 파일과 출력 파일을 선택해주세요.")
        return
    
    try:
        name_col = int(name_col_entry.get())
        amount_col = int(amount_col_entry.get())
        phone_col = int(phone_col_entry.get())
        subject_col = int(subject_col_entry.get())
        discount_type_col = int(discount_type_col_entry.get())
        start_row = int(start_row_entry.get())

        prefix = prefix_entry.get()
        suffix = suffix_entry.get()
        message_text = message_entry.get("1.0", tk.END).strip()

        # 워크북 로드
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        # 데이터 저장용 딕셔너리
        data = defaultdict(lambda: {"total_amount": 0, "phone": "", "subjects": defaultdict(int), "discount_types": set()})

        # 이름, 금액, 전화번호, 과목, 할인유형을 읽음
        for row in ws.iter_rows(min_row=start_row):
            name = row[name_col - 1].value
            amount = row[amount_col - 1].value
            phone = row[phone_col - 1].value
            subject_info = row[subject_col - 1].value
            discount_type = row[discount_type_col - 1].value

            if amount is None :
                break

            # 이름을 기준으로 금액 합산
            data[name]["total_amount"] += amount
            data[name]["phone"] = phone

            # 과목 처리
            subjects = ["수학", "영어", "독서", "국어", "과학", "사회"]
            for subject in subjects:
                if subject in subject_info:
                    data[name]["subjects"][subject] += amount

            # 할인유형 처리
            if discount_type:
                data[name]["discount_types"].add(f"({discount_type} 적용)")

        # 결과 파일 생성
        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active

        # 제목 행 작성
        result_ws.append(["수취인", "전화번호", "청구금액", "청구사유", "안내메세지"])

        # 2행부터 데이터 입력
        for name, info in data.items():
            formatted_amount = f"{info['total_amount']:,}"  # 쉼표 추가된 금액 형식

            subject_details = ""
            for subject, subj_amount in info["subjects"].items():
                subject_details += f"[{subject}] {subj_amount:,} "

            discount_details = " ".join(info["discount_types"])

            claim_reason = f"{prefix}{formatted_amount}{suffix} {subject_details.strip()} {discount_details}".strip()
            result_ws.append([name, info["phone"], formatted_amount, claim_reason, message_text])

        # 결과 파일 저장
        result_wb.save(output_file)

        if os.path.exists(output_file):
            messagebox.showinfo("완료", f"결과 파일이 {output_file}에 저장되었습니다.")
        else:
            messagebox.showerror("오류", "결과 파일을 생성하지 못했습니다.")
    except Exception as e:
        messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다: {str(e)}")

# GUI 생성
root = tk.Tk()
root.title("BillConv_Ho")

# 열 선택 및 행 선택 (왼쪽에 나열)
left_frame = tk.Frame(root)
left_frame.grid(row=0, column=0, padx=10, pady=10)

name_col_label = tk.Label(left_frame, text="이름 열 번호:")
name_col_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")
name_col_entry = tk.Entry(left_frame, width=5)
name_col_entry.insert(0, "2")  # 기본값 2
name_col_entry.grid(row=0, column=1, padx=5, pady=5)

amount_col_label = tk.Label(left_frame, text="금액 열 번호:")
amount_col_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
amount_col_entry = tk.Entry(left_frame, width=5)
amount_col_entry.insert(0, "11")  # 기본값 11
amount_col_entry.grid(row=1, column=1, padx=5, pady=5)

phone_col_label = tk.Label(left_frame, text="전화번호 열 번호:")
phone_col_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
phone_col_entry = tk.Entry(left_frame, width=5)
phone_col_entry.insert(0, "6")  # 기본값 6
phone_col_entry.grid(row=2, column=1, padx=5, pady=5)

subject_col_label = tk.Label(left_frame, text="과목 열 번호:")
subject_col_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
subject_col_entry = tk.Entry(left_frame, width=5)
subject_col_entry.insert(0, "8")  # 기본값 8
subject_col_entry.grid(row=3, column=1, padx=5, pady=5)

discount_type_col_label = tk.Label(left_frame, text="할인유형 열 번호:")
discount_type_col_label.grid(row=4, column=0, padx=5, pady=5, sticky="e")
discount_type_col_entry = tk.Entry(left_frame, width=5)
discount_type_col_entry.insert(0, "15")  # 기본값 15
discount_type_col_entry.grid(row=4, column=1, padx=5, pady=5)

start_row_label = tk.Label(left_frame, text="시작 행 번호:")
start_row_label.grid(row=5, column=0, padx=5, pady=5, sticky="e")
start_row_entry = tk.Entry(left_frame, width=5)
start_row_entry.insert(0, "5")  # 기본값 5
start_row_entry.grid(row=5, column=1, padx=5, pady=5)

# 청구사유 텍스트 및 안내 메세지 (오른쪽에 배치)
right_frame = tk.Frame(root)
right_frame.grid(row=0, column=1, padx=10, pady=10)

prefix_label = tk.Label(right_frame, text="청구사유 앞 문구:")
prefix_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
prefix_entry = tk.Entry(right_frame, width=40)
prefix_entry.grid(row=0, column=1, padx=5, pady=5)

suffix_label = tk.Label(right_frame, text="청구사유 뒤 문구:")
suffix_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
suffix_entry = tk.Entry(right_frame, width=40)
suffix_entry.grid(row=1, column=1, padx=5, pady=5)

message_label = tk.Label(right_frame, text="안내 메세지:")
message_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
message_entry = tk.Text(right_frame, height=4, width=40)
message_entry.grid(row=2, column=1, padx=5, pady=5)

# 처리 버튼
process_button = tk.Button(root, text="변환하기", command=process_files, font=("Helvetica", 14, "bold"), bg="lightblue", width=20, height=2)
process_button.grid(row=1, column=0, columnspan=2, pady=10)

# 파일 선택 버튼 및 경로 표시 (아래쪽에 배치)
file_frame = tk.Frame(root)
file_frame.grid(row=2, column=0, columnspan=2, pady=10)

input_file_button = tk.Button(file_frame, text="파일 선택 (통통통)", command=select_input_file)
input_file_button.grid(row=0, column=0, padx=10, pady=10, sticky="w")

input_file_label = tk.Label(file_frame, text=": 없음", width=50, anchor="w")
input_file_label.grid(row=0, column=1, padx=10, pady=10, columnspan=2)

output_file_button = tk.Button(file_frame, text="저장 위치 (결제선생)", command=select_output_file)
output_file_button.grid(row=1, column=0, padx=10, pady=10, sticky="w")

output_file_label = tk.Label(file_frame, text=": 없음", width=50, anchor="w")
output_file_label.grid(row=1, column=1, padx=10, pady=10, columnspan=2)

root.mainloop()
